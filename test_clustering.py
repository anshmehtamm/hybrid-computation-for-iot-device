from heuristic_clustering import *
import graph_clustering
import distance_clustering
from setupNodes import *
import threading 
import datetime
from server import Server
from client import Client
from distributed_server import dist_Server
from distributed_client import dist_Client
from node import Node
from openpyxl import Workbook
import multiprocessing
import time
import os
from queue import Queue
import generateEvents


que = Queue()

def run_server(node):
    sx = Server(node)
    sx.server_script(que)

def run_client(node):
    sx = Client(node)
    sx.client_script(que)

def run_d_server(node):
    sx = dist_Server(node)
    sx.server_script(que)


def run_d_client(node):
    sx = dist_Client(node)
    sx.client_script()

def write_thread1(dirr):
    wb = Workbook()
    ws = wb.active
    ws.cell(column=1, row=1).value = "Mode of Communication"
    ws.cell(column=2, row=1).value = "Cluster ID"
    ws.cell(column=3, row=1).value = "Event ID"
    ws.cell(column=4, row=1).value = "SRC Node ID"
    ws.cell(column=5, row=1).value = "DEST Node ID"
    ws.cell(column=6, row=1).value = "Timestamp"
    ws.cell(column=7, row=1).value = "Microsecond"
    r = 2
    print("Writing thread started")
    xx = os.path.join(dirr,'Communication Recordings.xlsx')
    while True:
        worker = que.get()
        #print("GOT")
        for h in worker:
            for i in range(len(h)):
                ws.cell(column=i+1, row=r).value = h[i]
            r+=1
        wb.save(xx)
        print(r,end=" ")

def write_thread2(dirr):
    wb = Workbook()
    ws = wb.active
    ws.cell(column=1, row=1).value = "Event ID"
    ws.cell(column=2, row=1).value = "Cluster ID"
    ws.cell(column=3, row=1).value = "Type of Cluster"
    ws.cell(column=4, row=1).value = "Propogation Time"

    r = 2
    print("Writing thread started")
    xx = os.path.join(dirr,'EventPropogationTime.xlsx')
    
    while True:
        worker = que.get()
        #print("GOT")
        i = 1
       # print(worker)
        for h in worker:
        #    print(h)
            ws.cell(column=i, row=r).value = h
            i+=1
        r+=1
        wb.save(xx)
        print(r,end=" ")


def start_comm(nodesList):
    for n in nodesList:
        if n.cluster_head==True:
            if n.computing==1:
                t1 = threading.Thread(target=run_server, args=((n),))
                t1.start()
                t3 = threading.Thread(target=run_client,args=((n),))
                t3.start()
                for ne in n.neighbor_tran_range:
                    if ne.cluster_id==n.cluster_id:
                        t2 = threading.Thread(target=run_client,args=((ne),))
                        t2.start()
            else:
                t1 = threading.Thread(target=run_d_server, args=((n),))
                t3 = threading.Thread(target=run_d_client,args=((n),))
                t3.start()
                for ne in n.neighbor_tran_range:
                    if ne.cluster_id==n.cluster_id:
                        t2 = threading.Thread(target=run_d_client,args=((ne),))
                        t2.start()
                t1.start()


def set_comm_parameters(nodesList):

    cluster_head_count = 0
    cluster_node_count = {}
    cl = {}

    for node in nodesList:
        if node.cluster_id not in cl:
            cl[node.cluster_id]=[]
        if node.cluster_head==True:
            cluster_head_count+=1
        cl[node.cluster_id].append(node)
        cluster_node_count[node.cluster_id] = cluster_node_count.get(node.cluster_id,0)+1
    lis = sorted(cluster_node_count.items(),key=lambda x: -x[1])
    cs,d=0,0
    print(len(lis))
    x = 0
    for cl_id,count in lis:
        com = 0
        if len(cl[cl_id])!=0:
            if len(cl[cl_id])>7:
                d+=1
                com=0
            else:
                cs+=1
                com=1
        else:
            x+=1
            continue

        for n in cl[cl_id]:
            n.computing=com
            if n.cluster_head==True:
                if com==1:
                    t = plt.annotate('Client-Server', xy=(n.grid_x, n.grid_y), xytext=(n.grid_x+10, n.grid_y+10),arrowprops=dict(facecolor='black', shrink=0.0001))
                  #  plt.pause(0.5)
                    t.set_visible(False)
                else:
                    t = plt.annotate('Distributed', xy=(n.grid_x, n.grid_y), xytext=(n.grid_x+10, n.grid_y+10),arrowprops=dict(facecolor='black', shrink=0.0001))
                   # plt.pause(0.5)
                    t.set_visible(False)   
    #fig = plt.figure()
    alone=0
    for node in nodesList:
        if len(node.neighbor_tran_range)==0:
            alone+=1
    cs-=alone
    plt.suptitle('Client-Server = %d, Distributed = %d'%(cs,d), fontsize=14, fontweight='bold')

def store_node_info(nodesList1,dirr):
    wb = Workbook()
    ws = wb.active
    ws.cell(column=1, row=1).value = "Node ID"
    ws.cell(column=2, row=1).value = "Position"
    ws.cell(column=3, row=1).value = "Transmission Range"
    ws.cell(column=4, row=1).value = "Neighbors Count"
    ws.cell(column=5, row=1).value = "Neighbors Node ID"
    ws.cell(column=6, row=1).value = "Alone Node"
    ws.cell(column=7, row=1).value = "Cluster ID"
    ws.cell(column=8, row=1).value = "Cluster Head"
    ws.cell(column=9,row=1).value = "Computing"
    
    r = 2
    for node in nodesList1:
        
        ws.cell(column=1,row=r).value = node.nodeId
        x = ""
        x+=str(node.grid_x)
        x+=" "
        x+=str(node.grid_y)
        ws.cell(column=2,row=r).value = x
        ws.cell(column=3, row=r).value = node.tran_range
        ws.cell(column=4, row=r).value = len(node.neighbor_tran_range)
        x = ""
        for n in node.neighbor_tran_range:
            x+=str(n.nodeId)
            x+=", "
        ws.cell(column=5, row=r).value = x[:len(x)-2]
        if len(node.neighbor_tran_range)==0:
            ws.cell(column=6, row=r).value = "True"
        else:
            ws.cell(column=6, row=r).value = "False"

        ws.cell(column=7, row=r).value = node.cluster_id

        ws.cell(column=8, row=r).value = node.cluster_head
        if node.computing==0:
            ws.cell(column=9,row=r).value = "Distributed"
        else:
            ws.cell(column=9,row=r).value = "Client/Server"
        r+=1
    xx = os.path.join(dirr,'Detailed Info.xlsx')
    wb.save(xx)
            
def store_dep_info(nodesList1,nn,ha,dirr,cl,al):
    n_c = len(nodesList1)
    numb_clust,dist,cl_s,alone = 0,0,0,0
    kk =set()
    for n in nodesList1:
        if n.cluster_id not in kk:
            if n.computing==0:
                dist+=1
            else:
                cl_s+=1
    wb = Workbook()
    ws = wb.active
    ws.cell(column=1, row=1).value = "Total Nodes Deployed"
    ws.cell(column=2, row=1).value = n_c

    ws.cell(column=1, row=2).value = "Deployed Randomly"
    ws.cell(column=2, row=2).value = nn

    ws.cell(column=1, row=3).value = "Deployed using Halton"
    ws.cell(column=2, row=3).value = ha

    ws.cell(column=1, row=4).value = "Number of Clusters"
    ws.cell(column=2, row=4).value = cl

    ws.cell(column=1, row=5).value = "Alone Nodes"
    ws.cell(column=2, row=5).value = al

    ws.cell(column=1, row=6).value = "Distributed"
    ws.cell(column=2, row=6).value = dist

    ws.cell(column=1, row=7).value = "Client/Server"
    ws.cell(column=2, row=7).value = cl_s

    xx = os.path.join(dirr,'Info.xlsx')
    wb.save(xx)

def plot_barplot(d_c,g_c,h_c,d_a,g_a,h_a,path):

    import pandas as pd

    data = {'Number of Clusters': {"K-Means":d_c,"Graph Based":g_c,"Heuristic Based":h_c}, 'Alone Nodes': {"K-Means":d_a, "Graph Based":g_a, "Heuristic Based":h_a}}
    df = pd.DataFrame(data)
    df.plot(kind='bar')
    plt.ylabel('Count')
    x = os.path.join(path,'Bar-Plot comparing Clustering Techniques')
    plt.savefig(x,bbox_inches='tight')
    plt.close()

def identify_neighbours(nodesList):
    for node in nodesList:
        neigh = []
        for m in nodesList:
            if m is node:
                continue
            if ((m.grid_x-node.grid_x)**2 + (m.grid_y-node.grid_y)**2) <= (node.tran_range)**2:
                neigh.append(m)
        node.setNeighbors(neigh) 


    



#n = int(input("Enter Number of Nodes in Se`nsing Layer"))
#ha = int(input("No of Nodes to be deployed using Halton Strategy"))
#grid_x,grid_y =# [int(x) for x in input("Enter Size of Grid (x y)").split()]
if __name__ == "__main__":
    

    parent_dir = r"D:\Test"
    todo = [[300,0,100,100,10]]

    for i in range(len(todo)):

        n =todo[i][0]
        ha = todo[i][1]
        tr_d = todo[i][4]
        grid_x,grid_y = todo[i][2],todo[i][3]

        directory = "(%d - %d, %d, %d,%d)"%(i,n,ha,grid_x,grid_y)
        path = os.path.join(parent_dir, directory)
        os.mkdir(path)

        nodesList1 = form_nodes(n+ha,n,ha)

        form_grid_random(nodesList1[:n],grid_x,grid_y)
        form_grid_haldon(nodesList1[n:],grid_x,grid_y,2,3)

        set_tran_range(nodesList1,tr_d)
        set_re(nodesList1)

        # fig, ax = plt.subplots()
        # plt.scatter([node.grid_x for node in nodesList1[:n]],[node.grid_y for node in nodesList1[:n]],color='y',label="Random Nodes")
        # plt.title("Deployed Nodes")
        # plt.scatter([node.grid_x for node in nodesList1[n:]],[node.grid_y for node in nodesList1[n:]],color='b',label="Halton Nodes")
        # print(len(nodesList1))
        # plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', borderaxespad=0.)
        # plt.xlabel('X - axis')

        # plt.ylabel('Y - axis')
        # plt.figtext(0.5, -0.05, "Number of Nodes = %d, Grid Size = %d x %d, Transmission Dist = %d"%(len(nodesList1),grid_x,grid_y,tr_d), ha="center")
        # x = os.path.join(path,'Deployed Nodes')
        # plt.savefig(x,bbox_inches='tight')
        #plt.clf()

        identify_neighbours(nodesList1)
        # fig, ax = plt.subplots()
        # plt.scatter([node.grid_x for node in nodesList1],[node.grid_y for node in nodesList1],color='r',label="Node")
        # for node in nodesList1:
        #     for nn in node.neighbor_tran_range:
        #         plt.plot([node.grid_x,nn.grid_x],[node.grid_y,nn.grid_y],'k-',color='b')
        
        # plt.title("Connectivity of Deployement")
        # plt.xlabel("X - axis")
        # plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', borderaxespad=0.)
        # plt.ylabel("Y - axis")
        # x = os.path.join(path,'Connectivity of Deployment')
        # plt.figtext(0.5, -0.05, "Number of Nodes = %d, Grid Size = %d x %d, Transmission Dist = %d "%(len(nodesList1),grid_x,grid_y,tr_d), ha="center")
        # plt.savefig(x,bbox_inches='tight')
        # plot_deployed(path,nodesList1,grid_x,grid_y,n,ha,tr_d)
        #plt.show()
        #plt.clf()
        nodesList2 = form_copy(nodesList1)
        nodesList3 = form_copy(nodesList1)
        identify_neighbours(nodesList2)
        identify_neighbours(nodesList3)

        dist_dir = os.path.join(path,"Distance-Based Clustering")
        os.mkdir(dist_dir)
        x =  os.path.join(dist_dir,"KMeans Clustering")

        graph_dir =  os.path.join(path,'Graph-Based Clustering')
        os.mkdir(graph_dir)
        y =  os.path.join(graph_dir,"Graph Clustering")

        heir_dir = os.path.join(path,'Heuristic-Based Clustering')
        os.mkdir(heir_dir)
        z = os.path.join(heir_dir,"Heuristic Clustering")

        #d_c,d_a = distance_clustering.k_means(nodesList2,grid_x,grid_y)
        #plt.savefig(x,bbox_inches='tight')


        g_c,g_a = graph_clustering.graph_clustering(nodesList1,grid_x,grid_y)

        #plt.savefig(y,bbox_inches='tight')

        #h_c,h_a = heuristic_clustering(nodesList3,grid_x,grid_y)
        #plt.savefig(z,bbox_inches='tight')

        set_comm_parameters(nodesList1)
        set_comm_parameters(nodesList2)
        set_comm_parameters(nodesList3)

       # store_node_info(nodesList2,dist_dir)
       #store_dep_info(nodesList2,n,ha,dist_dir,g_c,g_a)

        store_node_info(nodesList1,graph_dir)
        store_dep_info(nodesList1,n,ha,graph_dir,g_c,g_a)

        #store_node_info(nodesList3,heir_dir)
        #store_dep_info(nodesList3,n,ha,heir_dir,h_c,h_a)

        #plot_barplot(d_c,g_c,h_c,d_a,g_a,h_a,path)


        wr_t = threading.Thread(target=write_thread2,args=((path,)))
        wr_t.start()

        ev_t = threading.Thread(target=generateEvents.generate,args=((nodesList1,grid_x,grid_y,parent_dir,)))
        ev_t.start()
        start_comm(nodesList1)
        
        
        wr_t.join()



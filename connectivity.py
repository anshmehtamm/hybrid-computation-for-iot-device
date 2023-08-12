from setupNodes import *
from openpyxl import Workbook
from math import comb
import os
import networkx as nx

PARENT_DIR = "D:/Test"


def setupWorkBook():
    wb = Workbook()
    ws = wb.active

    ws.cell(column=1, row=1).value = "Technique"
    ws.cell(column=2, row=1).value = "Random Nodes"
    ws.cell(column=3, row=1).value = "Halton Nodes"
    ws.cell(column=4, row=1).value = "Total Nodes"
    ws.cell(column=5, row=1).value = "Average Connectivity"
    ws.cell(column=6, row=1).value = "Average Reachibility"
    
    return wb,ws

def save(detailList,ws,r):

    for i,it in enumerate(detailList):
        ws.cell(column=i+1, row=r).value = it

    return r+1

def create_graph(nodesList1):

    G = nx.DiGraph()

    for n in nodesList1:
        for x in n.neighbor_tran_range:
            G.add_edge(n.nodeId,x.nodeId)
    
    return G






wb,ws = setupWorkBook()

t = [[50,0],[100,0],[150,0],[200,0],[250,0],[300,0],[350,0],[400,0],[450,0],[500,0],[0,50],[0,100],[0,150],[0,200],[0,250],[0,300],[0,350],[0,400],[0,450],[0,500],[37,13],[75,25],[112,38],[150,50],[187,63],[225,75],[262,88],[300,100],[337,113],[375,125]]

r = 2

for i in range(len(t)):

    avgc = 0
    avgr = 0

    for j in range(10):

        n,ha,grid_x,grid_y = t[i][0],t[i][1],100,100


        nodesList1 = form_nodes(n+ha,n,ha)

        form_grid_random(nodesList1[:n],grid_x,grid_y)
        form_grid_haldon(nodesList1[n:],grid_x,grid_y,2,3)
        set_tran_range(nodesList1,10)

        identifyNeighbours(nodesList1)

        G = create_graph(nodesList1)

        conn = 0
        for no in nodesList1:
            conn+=no.nc
            #print(no.nc)
        conn = conn/(2*comb(n+ha,2))
       # print(conn)
        avgc+=conn

        reach = 0
        sub_graphs = nx.strongly_connected_components(G)

        for x in sub_graphs:
            reach += comb(len(x),2)
        reach = reach/comb(n+ha,2)
        avgr+=reach


    detailList = None

    if ha==0:
        detailList = ["Random",n,ha,n+ha,avgc/10,avgr/10]
        
    elif n==0:
        detailList = ["Halton",n,ha,n+ha,avgc/10,avgr/10]

    else:
        detailList = ["Random+Halton",n,ha,n+ha,avgc/10,avgr/10]
    
    r = save(detailList,ws,r)

name = "ConnectivityReport.xlsx"
PATH = os.path.join(PARENT_DIR, name)

wb.save(PATH)


    









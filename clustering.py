from setupNodes import *
from openpyxl import Workbook
from math import comb
import os
import graph_clustering
import distance_clustering
import heuristic_clustering
import generateEvents
import networkx as nx

PARENT_DIR = "D:/Test"


def setupWorkBook():
    wb = Workbook()
    ws = wb.active

    ws.cell(column=1, row=1).value = "Technique"
    ws.cell(column=2, row=1).value = "Random Nodes"
    ws.cell(column=3, row=1).value = "Halton Nodes"
    ws.cell(column=4, row=1).value = "Total Nodes"
    ws.cell(column=5, row=1).value = "Clustering Technique"
    ws.cell(column=6, row=1).value = "Average Number of Clusters"
    ws.cell(column=7, row=1).value = "Average Alone Nodes"
    
    
    return wb,ws

def save(detailList,ws,r):

    for i,it in enumerate(detailList):
        ws.cell(column=i+1, row=r).value = it

    return r+1


wb,ws = setupWorkBook()

t = [[37,13,'G'],[75,25,'G'],[112,38,'G'],[150,50,'G'],[187,63,'G'],[225,75,'G'],[262,88,'G'],[300,100,'G'],[337,113,'G'],[375,125,'G'],[37,13,'H'],[75,25,'H'],[112,38,'H'],[150,50,'H'],[187,63,'H'],[225,75,'H'],[262,88,'H'],[300,100,'H'],[337,113,'H'],[375,125,'H'],[37,13,'D'],[75,25,'D'],[112,38,'D'],[150,50,'D'],[187,63,'D'],[225,75,'D'],[262,88,'D'],[300,100,'D'],[337,113,'D'],[375,125,'D']]

r = 2

for i in range(len(t)):

    avgalone = 0
    avgclusters = 0
    ty = ""

    for j in range(10):

        n,ha,grid_x,grid_y = t[i][0],t[i][1],100,100


        nodesList1 = form_nodes(n+ha,n,ha)

        form_grid_random(nodesList1[:n],grid_x,grid_y)
        form_grid_haldon(nodesList1[n:],grid_x,grid_y,2,3)

        set_tran_range(nodesList1,10)
        identifyNeighbours(nodesList1)

        x,y = 0,0

        if t[i][2]=='G':
            x,y = graph_clustering.graph_clustering(nodesList1,grid_x,grid_y)
            ty = "Graph Based"
        elif t[i][2]=='H':
            x,y= heuristic_clustering.heuristic_clustering(nodesList1,grid_x,grid_y)
            ty = "Heuristic Based"
        else:
            x,y = distance_clustering.k_means(nodesList1,grid_x,grid_y)
            ty = "Distance  Basedd"
    
        avgalone += y
        avgclusters += x

    detailList = None

    if ha==0:
        detailList = ["Random",n,ha,n+ha,ty,avgclusters/10,avgalone/10]
        
    elif n==0:
        detailList = ["Halton",n,ha,n+ha,ty,avgclusters/10,avgalone/10]

    else:
        detailList = ["Random+Halton",n,ha,n+ha,ty,avgclusters/10,avgalone/10]
    
    r = save(detailList,ws,r)

name = "ClusteringReport.xlsx"
PATH = os.path.join(PARENT_DIR, name)

wb.save(PATH)

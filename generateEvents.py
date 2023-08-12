
import random
import datetime  
import event
import node
import socket
from openpyxl import Workbook
import pickle
import os
import threading 

def generate(nodesList,grid_x,grid_y,parent_dir):
    
    name = "EventDetails.xlsx"
    PATH = os.path.join(parent_dir, name)

    #wb,ws = setupWriteBook()
    ROW_NO = 2

    numberOfEvents = 50
    totalDropped = 0

    for ev_no in range(numberOfEvents):

        x = random.randint(1,grid_x)
        y = random.randint(1,grid_y)
        time = datetime.datetime.now() 
        
        ev = event.event(ev_no+1,x,y,time)
       # print("Event %d at %d,%d"%(ev_no+1,x,y))

        totalDropped += notifyAll(ev,nodesList)
        #ROW_NO = saveEvent(ws,ROW_NO,ev)
    return totalDropped
    #wb.save(PATH)    


def notifyAll(event,nodesList):

    euclideanThreshold = 5
    notifiedCount = 0
    sensed_by = []
    dropped = 0
    for n in nodesList:
        if (((n.grid_x-event.x)**2 + (n.grid_y-event.y)**2)**0.5) <= euclideanThreshold:
            
            sensed_by.append(n.nodeId)

            #send_t = threading.Thread(target=send,args=((event,n.nodeId,)))
            #send_t.start()
            
            send(event,n.nodeId)
            #print("sent %d to node %d"%(event.ev_id,n.nodeId))

            notifiedCount+=1
            
    print("Notified to : ",notifiedCount)

    if notifiedCount==0:
        #no one sensed this event add to report...
        dropped+=1

     #   print("No one nearby")
        
    event.sensed_by = sensed_by
    return dropped




def send(event,id):
    host = '127.0.0.1'  
    
    #PORT BASED ON TYPE OF COMM
    port = id+16000

    s = socket.socket(socket.AF_INET,socket.SOCK_STREAM)
    
    try:
        s.connect((host,port))
        s.send(pickle.dumps(event))
        s.close()
    except:
        print("Error in sending Event%d at %d "%(event.ev_id,id))  

        
def setupWriteBook():

    wb = Workbook()
    ws = wb.active
    ws.cell(column=1, row=1).value = "Event ID"
    ws.cell(column=2, row=1).value = "Position X"
    ws.cell(column=3, row=1).value = "Position Y"
    ws.cell(column=4,row=1).value = "Time of Occurence"
    ws.cell(column=5,row=1).value = "Sensed By"

    return wb,ws

def saveEvent(ws,r,ev):

    detailList = ev.getList()

    for i,it in enumerate(detailList):
        ws.cell(column=i+1, row=r).value = it
    
    ws.cell(column=5,row=r).value = ev.getSensed()
    return r+1



        






